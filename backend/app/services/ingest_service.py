"""
Ingest service — Excel → book records.
Ported from input_handler.py.
"""

import structlog
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.services.book_service import create_book

logger = structlog.get_logger(__name__)


async def ingest_excel(
    db: AsyncSession,
    user_id: str,
    file_bytes: bytes,
    selected_model: str | None = None,
) -> list[str]:
    """
    Parse an Excel file and create book records for each row.

    Expected columns: title (required), notes_on_outline_before (optional)
    Returns list of created book IDs.
    """
    import io
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel file is empty")

    headers = [str(c).strip().lower() if c else "" for c in rows[0]]

    if "title" not in headers:
        raise ValueError("Excel file must have a 'title' column")

    title_idx = headers.index("title")
    notes_idx = headers.index("notes_on_outline_before") if "notes_on_outline_before" in headers else None

    created_ids: list[str] = []

    for row in rows[1:]:
        title = str(row[title_idx]).strip() if row[title_idx] else ""
        if not title or title.lower() == "none":
            continue

        notes = ""
        if notes_idx is not None and row[notes_idx]:
            notes = str(row[notes_idx]).strip()

        book = await create_book(
            db=db,
            user_id=user_id,
            title=title,
            selected_model=selected_model,
        )
        # Store notes in outline_raw temporarily until outline_service picks it up
        book.outline_raw = notes
        db.add(book)

        created_ids.append(book.id)

    wb.close()
    logger.info("excel_ingested", count=len(created_ids), user_id=user_id)
    return created_ids
